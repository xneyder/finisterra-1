import click
import os
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.comments import Comment

@click.command()
@click.option('--output-file', '-o', required=True, help='Path to the output Excel file.')
@click.option('--sheet-name', '-s', required=True, help='Name of the sheet to create or replace.')
@click.argument('json-file', type=click.Path(exists=True))
def main(output_file, json_file, sheet_name):
    # Load the input JSON file to a dictionary
    with open(json_file) as f:
        data = json.load(f)

    # Extract the relevant fields from the resources list
    rows = []
    for resource in data['resources']:
        resource_type = resource['type']
        for instance in resource['instances']:
            instance_id = instance['attributes']['id']
            attributes = instance['attributes']
            row = [resource_type, instance_id, json.dumps(attributes)]
            rows.append(row)
    
    if not rows:
        print(f'No resources found in {json_file}')
        return

    # Load the existing workbook or create a new one
    if os.path.exists(output_file):
        workbook = load_workbook(output_file)
    else:
        workbook = Workbook()

        # Remove the default 'Sheet' when creating a new workbook
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

    # If the sheet with the specified name exists, use it. Otherwise, create a new sheet.
    if sheet_name in workbook:
        sheet = workbook[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(sheet_name)

    # Set the header row bold
    header_font = Font(bold=True)
    sheet.append(['Resource Type', 'ID', 'Attributes'])
    for cell in sheet[1]:
        cell.font = header_font

    # Write the rows to the sheet
    for row in rows:
        sheet.append(row)

    # Set the popover for each id cell
    for i, row in enumerate(rows):
        id_cell = sheet.cell(i+2, 2)
        attributes = json.loads(row[2])
        popover = '\n'.join([f'{k}: {v}' for k, v in attributes.items()])
        comment = Comment(popover, 'Details')
        id_cell.comment = comment

    # Remove the attributes column
    sheet.delete_cols(3)

    # Save the workbook
    workbook.save(output_file)

if __name__ == '__main__':
    main()
